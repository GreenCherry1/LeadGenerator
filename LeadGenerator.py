from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
from selenium.webdriver.common.actions.wheel_input import ScrollOrigin
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl
import requests
import json


class Scraper:
    def __init__(self):
        self.driver = webdriver.Chrome()

    @staticmethod
    def scroll(self):
        """
        Perform scrolling action on the web page.

        Args:
            self.driver (webdriver.driver): The Selenium webdriver.driver instance.
        """
        scroll = self.driver.find_elements(By.CLASS_NAME, 'hfpxzc')[-1]
        scroll_origin = ScrollOrigin.from_element(scroll, 0, 0)
        ActionChains(self.driver) \
            .move_to_element(scroll) \
            .scroll_from_origin(scroll_origin, 0, 100) \
            .perform()

    @staticmethod
    def close_tab(self):
        """
        Close the current tab and switch to the last tab.

        Args:
            self.driver (webdriver.driver): The Selenium webdriver.driver instance.
        """
        self.driver.close()
        self.driver.switch_to.window(self.driver.window_handles[-1])

    def get_email_page(self, pages, has_phone):
        """
        Extract email addresses from web pages.

        Args:
            self.driver (webdriver.driver): The Selenium webdriver.driver instance.
            pages (list): List of URLs to visit.
            has_phone (bool): Flag indicating if phone number is already found.

        Returns:
            dict: Dictionary containing the email and phone number found on the page.
        """
        headers = {'User-Agent': 'Chrome/92.0.4515.107'}
        out = {'phone': '', 'email': ''}
        for page in pages:
            try:
                code = requests.get(page, headers=headers).status_code
            except Exception as ex_:
                print(ex_)
                continue
            if code != 200:
                continue
            self.driver.execute_script("window.open('%s', '_blank')" % page)
            self.driver.implicitly_wait(10)
            self.driver.switch_to.window(self.driver.window_handles[-1])
            links = self.driver.find_elements(By.TAG_NAME, 'a')
            for link in links:
                try:
                    if "facebook" in link.get_attribute('href'):
                        email = self.get_email_facebook(link.get_attribute('href'))
                        if email:
                            self.close_tab(self)
                            out['email'] = email
                            if has_phone:
                                return out
                except Exception as ex_:
                    print(ex_)
                    continue
                try:
                    if not has_phone and 'tel' in link.get_attribute("href"):
                        out['phone'] = link.get_attribute("href")[4:]
                        if out['email']:
                            self.close_tab(self)
                            return out
                    if 'mailto' in link.get_attribute("href"):
                        out['email'] = link.get_attribute("href")[7:]
                        if has_phone or out['phone']:
                            self.close_tab(self)
                            return out
                except Exception as ex_:
                    print(ex_)
            self.close_tab(self)
        return out

    @staticmethod
    def log_in_facebook(self):
        """
        Log in to Facebook.

        Args:
            self.driver (webdriver.driver): The Selenium webdriver.driver instance.
        """
        email_input = self.driver.find_elements(By.XPATH, "//input[@type='email']")
        if len(email_input):
            email_input = email_input[0]
        else:
            return
        password_input = self.driver.find_element(By.XPATH, "//input[@type='password']")
        email_input.send_keys("lmvkd2008@gmail.com")
        password_input.send_keys("#9sbTFm//$3#G4i\n")

    def get_email_facebook(self, page):
        """
        Extract email address from Facebook page.

        Args:
            self.driver (webdriver.driver): The Selenium webdriver.driver instance.
            page (str): URL of the Facebook page.

        Returns:
            str: The email address found on the page.
        """
        self.driver.execute_script("window.open('%s', '_blank')" % page)
        self.driver.implicitly_wait(10)
        self.driver.switch_to.window(self.driver.window_handles[-1])
        self.log_in_facebook(self)
        info = self.driver.find_elements(By.XPATH,
                                         "//span[@class='x193iq5w xeuugli x13faqbe x1vvkbs x1xmvt09 x1lliihq x1s928wv "
                                         "xhkezso x1gmr53x x1cpjm7i x1fgarty x1943h6x xudqn12 x3x7a5m x6prxxf xvq8zen "
                                         "xo1l8bm xzsf02u x1yc453h']")
        for information in info:
            text = information.text
            if '@' in text:
                self.close_tab(self)
                return text
        self.close_tab(self)
        return ''

    def run_scraper(self, search, least_leads):
        """
        Run the web scraper.

        Args:
            search (str): The search query.
            least_leads (int): The minimum number of leads to be scraped.
        """
        self.driver.get("https://www.google.com.ua/maps/search/" + search + '?hl=en')
        soc_networking_list = [
            "facebook", "instagram", "glovo", "t.me/", "tiktok", "twitter", "reddit", "tumblr", "flickr", "instagtam"
        ]
        work_book = openpyxl.Workbook()
        work_sheet = work_book.active
        full_leads = 0
        links, used_links = [], 0
        while full_leads < least_leads:
            self.scroll(self)
            used_links += len(links)
            WebDriverWait(self.driver, timeout=10).until(
                lambda d: len(d.find_elements(By.CLASS_NAME, 'hfpxzc')) > used_links
                )
            links = [business_page.get_attribute('href') for business_page in
                     self.driver.find_elements(By.CLASS_NAME, 'hfpxzc')[used_links:]]
            if not len(links):
                print('no more places')
                break
            for link in links:
                self.driver.execute_script("window.open('%s', '_blank')" % link)
                self.driver.implicitly_wait(10)
                self.driver.switch_to.window(self.driver.window_handles[-1])
                pictures_buttons = self.driver.find_elements(By.CLASS_NAME, 'CsEnBe')
                pictures_aria_label = [x.get_attribute("aria-label") for x in pictures_buttons]
                websites = list(set([x.get_attribute('href') for x in pictures_buttons if
                                     x.tag_name == 'a' and x.get_attribute('aria-label') != 'Claim this business']))
                own_websites = [website for website in websites if sum([int(socnetwork in website) for socnetwork in
                                                                        soc_networking_list]) == 0 and website[
                                                                                                       -3:] != 'pdf']
                if not len(websites):
                    self.close_tab(self)
                    continue
                has_phone = False
                for pic in pictures_aria_label:
                    if "Plus code" in pic:
                        work_sheet.cell(row=full_leads + 1, column=2).value = pic[11:]
                    if "Phone" in pic:
                        work_sheet.cell(row=full_leads + 1, column=4).value = pic[6:]
                        has_phone = True
                email = {'email': '', 'phone': ''}
                if len(own_websites):
                    email = self.get_email_page(own_websites, has_phone)
                    work_sheet.cell(row=full_leads + 1, column=3).value = ' '.join(own_websites)
                else:
                    work_sheet.cell(row=full_leads + 2, column=3).value = ' '.join(websites)
                    for website in websites:
                        if 'facebook' in website:
                            email['email'] = self.get_email_facebook(website)
                            break
                work_sheet.cell(row=full_leads + 1, column=1).value = self.driver.find_element(
                    By.XPATH, "//h1[@class='DUwDvf fontHeadlineLarge']").text
                if email['phone']:
                    work_sheet.cell(row=full_leads + 1, column=4).value = email['phone']
                    has_phone = True
                if has_phone and email['email']:
                    work_sheet.cell(row=full_leads + 1, column=5).value = email['email']
                    full_leads += 1
                print(email)
                self.close_tab(self)
                if full_leads == Least_Leads:
                    break
        work_book.save("Leads.xlsx")
        self.driver.quit()


with open("settings.json") as jsonfile:
    data = json.load(jsonfile)
    key_words = list(data["KeyWords"])
    for key_word_index in range(len(key_words)):
        key_words[key_word_index] = '+'.join(key_words[key_word_index].split())
    Search = '+'.join(data["Place"].split()) + '+' + '+'.join(key_words)
    Least_Leads = data["Amount"]
scraper = Scraper()
scraper.run_scraper(Search, Least_Leads)
